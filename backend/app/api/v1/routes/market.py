"""
Market-level summary endpoints.
Dashboard reads from pre-computed snapshot tables populated by
compute/engine/market_snapshot.py after each nightly universe build.
screener.universe is never queried here — snapshot tables own market-level data.
"""
import logging
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import text

log = logging.getLogger(__name__)

from app.db.session import get_db
from app.core.cache import cache_get, cache_set, make_key, MARKET_TTL, STATIC_TTL
from app.schemas.market import (
    MarketSummary,
    MoversResponse,
    MoverStock,
    SectorsResponse,
    SectorStat,
    MarketDashboard,
    IndexSnapshot,
    DashboardStock,
    ActiveStock,
    VolumePressureStock,
    SectorHeatmapItem,
    ExDivStock,
    VolumeActivityResponse,
    HeatmapRow,
    HeatmapResponse,
)

router = APIRouter()


@router.get("/summary", response_model=MarketSummary)
async def market_summary(db: AsyncSession = Depends(get_db)):
    """
    Aggregate stats for the entire ASX universe.
    Used for the homepage stats bar.
    """
    _key = make_key("market", "summary")
    cached = await cache_get(_key)
    if cached:
        return MarketSummary(**cached)

    sql = text("""
        SELECT
            COUNT(*)                                                                    AS total_stocks,
            COUNT(*) FILTER (WHERE is_asx200)                                          AS asx200_stocks,
            COUNT(*) FILTER (WHERE dividend_yield > 0)                                 AS stocks_with_dividends,
            AVG(dividend_yield) FILTER (WHERE dividend_yield > 0
                                          AND dividend_yield < 0.30)                   AS avg_dividend_yield,
            PERCENTILE_CONT(0.5) WITHIN GROUP (ORDER BY pe_ratio)
                FILTER (WHERE pe_ratio > 0 AND pe_ratio < 100)                         AS median_pe,
            SUM(market_cap) / 1000.0                                                   AS total_market_cap_bn,
            MAX(universe_built_at)                                                     AS universe_built_at
        FROM screener.universe
        WHERE status = 'active'
    """)
    row = (await db.execute(sql)).mappings().one()
    result = MarketSummary(
        total_stocks=int(row["total_stocks"] or 0),
        asx200_stocks=int(row["asx200_stocks"] or 0),
        stocks_with_dividends=int(row["stocks_with_dividends"] or 0),
        avg_dividend_yield=float(row["avg_dividend_yield"]) if row["avg_dividend_yield"] is not None else None,
        median_pe=float(row["median_pe"]) if row["median_pe"] is not None else None,
        total_market_cap_bn=float(row["total_market_cap_bn"]) if row["total_market_cap_bn"] is not None else None,
        universe_built_at=row["universe_built_at"],
    )
    await cache_set(_key, result.model_dump(), ttl=MARKET_TTL)
    return result


@router.get("/movers", response_model=MoversResponse)
async def market_movers(
    period:   str           = Query("1w", pattern="^(1d|1w|1m|3m)$"),
    cap_tier: str | None    = Query(None, pattern="^(mega|large|mid|small|micro|nano|asx300)$"),
    limit:    int           = Query(10, ge=5, le=25),
    db: AsyncSession = Depends(get_db),
):
    """
    Top gainers and losers for a given period (1d / 1w / 1m / 3m).
    Optional cap_tier filter: large | mid | small | micro (uses market_cap ranges).
    asx300 = market_cap >= 300 (AUD millions), i.e. ASX 300+ universe.
    Without cap_tier, uses pre-computed mover_snapshots.
    """
    col_map = {"1d": "return_1d", "1w": "return_1w", "1m": "return_1m", "3m": "return_3m"}
    ret_col = col_map[period]

    # Period high/low column names in market.period_metrics
    ph_col = {"1d": "high_1d", "1w": "high_1w", "1m": "high_1m", "3m": "high_3m"}[period]
    pl_col = {"1d": "low_1d",  "1w": "low_1w",  "1m": "low_1m",  "3m": "low_3m"}[period]

    # LEFT JOIN to period_metrics using MAX(computed_date) — never CURRENT_DATE.
    # Using CURRENT_DATE races against the nightly compute job: if period_metrics
    # hasn't been written yet today the JOIN returns nothing and H/L show as dashes.
    # MAX(computed_date) always finds the most-recently-available data regardless of
    # when in the day the request arrives.
    pm_date_subq = "(SELECT MAX(computed_date) FROM market.period_metrics)"
    pm_join   = f"""LEFT JOIN market.period_metrics pm
                       ON pm.asx_code = u.asx_code
                      AND pm.computed_date = {pm_date_subq}"""
    pm_select = f"pm.{ph_col} AS period_high, pm.{pl_col} AS period_low"

    # Cap tier filter — uses pre-computed boolean flags (is_mega, is_large, etc.)
    # populated by build_screener_universe.py from raw AUD market_cap thresholds.
    # asx300 = direct market_cap >= 300M filter (all tiers ≥$300M).
    cap_flag_sql = {
        "mega":   "AND is_mega  = TRUE",
        "large":  "AND is_large = TRUE",
        "mid":    "AND is_mid   = TRUE",
        "small":  "AND is_small = TRUE",
        "micro":  "AND is_micro = TRUE",
        "nano":   "AND is_nano  = TRUE",
        "asx300": "AND market_cap >= 300",   # ≥$300M — excludes micro/nano noise
    }

    gainers_rows: list = []
    losers_rows:  list = []

    # ── Cap-tier filter: query screener.universe with market_cap ranges ───────
    if cap_tier:
        mcap_filter = cap_flag_sql[cap_tier]
        # Try requested period column first; fall back to return_1w if column missing
        for col in [ret_col, "return_1w"]:
            try:
                g_rows = (await db.execute(text(f"""
                    SELECT u.asx_code, u.company_name, u.sector,
                           u.price, u.{col} AS period_return, u.market_cap,
                           {pm_select}
                    FROM screener.universe u
                    {pm_join}
                    WHERE u.{col} IS NOT NULL
                      AND u.price > 0.05
                      AND u.market_cap > 0
                      {mcap_filter}
                    ORDER BY u.{col} DESC NULLS LAST
                    LIMIT :lim
                """), {"lim": limit})).mappings().all()
                l_rows = (await db.execute(text(f"""
                    SELECT u.asx_code, u.company_name, u.sector,
                           u.price, u.{col} AS period_return, u.market_cap,
                           {pm_select}
                    FROM screener.universe u
                    {pm_join}
                    WHERE u.{col} IS NOT NULL
                      AND u.price > 0.05
                      AND u.market_cap > 0
                      {mcap_filter}
                    ORDER BY u.{col} ASC NULLS LAST
                    LIMIT :lim
                """), {"lim": limit})).mappings().all()
                gainers_rows = list(g_rows)
                losers_rows  = list(l_rows)
                break  # success — stop trying fallback columns
            except Exception as e:
                log.warning("cap_tier movers (%s, col=%s) failed: %s", cap_tier, col, e)
                gainers_rows = []
                losers_rows  = []
                try:
                    await db.rollback()
                except Exception:
                    pass
        return MoversResponse(
            gainers=[_to_mover(r, period) for r in gainers_rows],
            losers =[_to_mover(r, period) for r in losers_rows],
            period=period,
        )

    # ── No cap filter: use pre-computed mover_snapshots ───────────────────────
    period_upper    = period.upper()
    gainer_type     = f"GAINER_{period_upper}"
    loser_type      = f"LOSER_{period_upper}"
    fallback_gainer = "GAINER"
    fallback_loser  = "LOSER"

    try:
        # Use MAX without date restriction so old snapshots still work
        snap_date_row = (await db.execute(text("""
            SELECT MAX(snapshot_date) AS latest
            FROM market.mover_snapshots
        """))).mappings().one()
        snap_date = snap_date_row["latest"]
        if snap_date:
            snap_rows = (await db.execute(text("""
                SELECT snapshot_type, asx_code, company_name, sector,
                       price, return_1w AS period_return, market_cap,
                       period_high, period_low
                FROM market.mover_snapshots
                WHERE snapshot_date = :d
                  AND snapshot_type IN (:gt, :lt, :fg, :fl)
                ORDER BY snapshot_type, rank
            """), {"d": snap_date, "gt": gainer_type, "lt": loser_type,
                   "fg": fallback_gainer, "fl": fallback_loser})).mappings().all()

            gainers_rows = [r for r in snap_rows if r["snapshot_type"] == gainer_type] \
                        or [r for r in snap_rows if r["snapshot_type"] == fallback_gainer]
            losers_rows  = [r for r in snap_rows if r["snapshot_type"] == loser_type]  \
                        or [r for r in snap_rows if r["snapshot_type"] == fallback_loser]
            log.info("movers snapshot=%s period=%s gainers=%d losers=%d",
                     snap_date, period, len(gainers_rows), len(losers_rows))
    except Exception as e:
        log.warning("mover_snapshots query failed: %s", e)

    # Snapshot rows now carry period_high / period_low directly (written by
    # market_snapshot.py since migration 052). No runtime enrichment needed.

    # Fallback: live universe query if snapshots empty
    if not gainers_rows and not losers_rows:
        # Try the requested period col; if missing (e.g. return_1d), try return_1w
        for col in [ret_col, "return_1w"]:
            try:
                g_rows = (await db.execute(text(f"""
                    SELECT u.asx_code, u.company_name, u.sector,
                           u.price, u.{col} AS period_return, u.market_cap,
                           {pm_select}
                    FROM screener.universe u
                    {pm_join}
                    WHERE u.{col} IS NOT NULL AND u.price > 0.05 AND u.market_cap > 20
                    ORDER BY u.{col} DESC NULLS LAST LIMIT :lim
                """), {"lim": limit})).mappings().all()
                l_rows = (await db.execute(text(f"""
                    SELECT u.asx_code, u.company_name, u.sector,
                           u.price, u.{col} AS period_return, u.market_cap,
                           {pm_select}
                    FROM screener.universe u
                    {pm_join}
                    WHERE u.{col} IS NOT NULL AND u.price > 0.05 AND u.market_cap > 20
                    ORDER BY u.{col} ASC NULLS LAST LIMIT :lim
                """), {"lim": limit})).mappings().all()
                gainers_rows = list(g_rows)
                losers_rows  = list(l_rows)
                log.info("movers universe fallback col=%s gainers=%d losers=%d",
                         col, len(gainers_rows), len(losers_rows))
                break
            except Exception as e:
                log.warning("screener.universe movers (col=%s) failed: %s", col, e)
                gainers_rows = []
                losers_rows  = []
                try:
                    await db.rollback()
                except Exception:
                    pass

    return MoversResponse(
        gainers=[_to_mover(r, period) for r in gainers_rows],
        losers =[_to_mover(r, period) for r in losers_rows],
        period=period,
    )


def _to_mover(r, period: str) -> MoverStock:
    """Map a DB row (with period_return alias) to MoverStock schema.

    period_high / period_low are always present in the row:
    - Snapshot path:        stored in market.mover_snapshots since migration 052
    - Cap-tier path:        LEFT JOIN to market.period_metrics in the SELECT
    - Live-fallback path:   LEFT JOIN to market.period_metrics in the SELECT
    All three paths use MAX(computed_date) so there is no CURRENT_DATE race condition.
    """
    pr = float(r["period_return"]) if r.get("period_return") is not None else None
    return MoverStock(
        asx_code=r["asx_code"],
        company_name=r["company_name"],
        sector=r.get("sector"),
        price=float(r["price"]) if r.get("price") is not None else None,
        return_1d=pr if period == "1d" else None,
        return_1w=pr if period == "1w" else None,
        return_1m=pr if period == "1m" else None,
        return_3m=pr if period == "3m" else None,
        market_cap=float(r["market_cap"]) if r.get("market_cap") is not None else None,
        period_high=float(r["period_high"]) if r.get("period_high") is not None else None,
        period_low =float(r["period_low"])  if r.get("period_low")  is not None else None,
    )


# ── helpers shared with volume-activity ──────────────────────────────────────

def _to_active(r) -> ActiveStock:
    return ActiveStock(
        asx_code=r["asx_code"],
        company_name=r["company_name"],
        sector=r.get("sector"),
        price=float(r["price"]) if r.get("price") is not None else None,
        return_1w=float(r["return_1w"]) if r.get("return_1w") is not None else None,
        market_cap=float(r["market_cap"]) if r.get("market_cap") is not None else None,
        volume=int(r["volume"]) if r.get("volume") is not None else None,
        avg_volume_20d=int(r["avg_volume_20d"]) if r.get("avg_volume_20d") is not None else None,
    )


def _to_vol_pressure(r) -> VolumePressureStock:
    vol = int(r["volume"]) if r.get("volume") is not None else None
    avg = int(r["avg_volume_20d"]) if r.get("avg_volume_20d") is not None else None
    ratio = (r["volume"] / r["avg_volume_20d"]) if (r.get("volume") and r.get("avg_volume_20d")) else None
    return VolumePressureStock(
        asx_code=r["asx_code"],
        company_name=r["company_name"],
        sector=r.get("sector"),
        price=float(r["price"]) if r.get("price") is not None else None,
        return_1w=float(r["return_1w"]) if r.get("return_1w") is not None else None,
        market_cap=float(r["market_cap"]) if r.get("market_cap") is not None else None,
        volume=vol,
        avg_volume_20d=avg,
        volume_ratio=float(ratio) if ratio is not None else None,
    )


@router.get("/volume-activity", response_model=VolumeActivityResponse)
async def volume_activity(
    cap_tier: str | None = Query(None, pattern="^(mega|large|mid|small|micro|nano|asx300)$"),
    limit:    int        = Query(10, ge=5, le=25),
    db: AsyncSession = Depends(get_db),
):
    """
    Most Active by Volume, Heavy Buying and Heavy Selling panels.
    Without cap_tier: uses pre-computed mover_snapshots (same as dashboard).
    With cap_tier: queries screener.universe live using boolean flag columns.
    asx300 = market_cap >= 300 (AUD millions).
    """
    cap_flag_sql = {
        "mega":   "AND is_mega  = TRUE",
        "large":  "AND is_large = TRUE",
        "mid":    "AND is_mid   = TRUE",
        "small":  "AND is_small = TRUE",
        "micro":  "AND is_micro = TRUE",
        "nano":   "AND is_nano  = TRUE",
        "asx300": "AND market_cap >= 300",
    }

    _key = make_key("market", "volume_activity", cap_tier or "all")
    cached = await cache_get(_key)
    if cached:
        return VolumeActivityResponse(**cached)

    active_rows:  list = []
    buying_rows:  list = []
    selling_rows: list = []

    if cap_tier:
        flag = cap_flag_sql[cap_tier]
        liquid = f"status = 'active' AND price > 0.05 AND market_cap > 0 {flag}"
        try:
            active_rows = (await db.execute(text(f"""
                SELECT asx_code, company_name, sector,
                       price, return_1w, market_cap, volume, avg_volume_20d
                FROM screener.universe
                WHERE {liquid}
                  AND volume IS NOT NULL AND volume > 0
                ORDER BY volume DESC NULLS LAST
                LIMIT :lim
            """), {"lim": limit})).mappings().all()

            buying_rows = (await db.execute(text(f"""
                SELECT asx_code, company_name, sector,
                       price, return_1w, market_cap, volume, avg_volume_20d
                FROM screener.universe
                WHERE {liquid}
                  AND volume IS NOT NULL AND avg_volume_20d IS NOT NULL AND avg_volume_20d > 0
                  AND return_1w > 0
                  AND volume::float / avg_volume_20d >= 1.5
                ORDER BY volume::float / avg_volume_20d DESC NULLS LAST
                LIMIT :lim
            """), {"lim": limit})).mappings().all()

            selling_rows = (await db.execute(text(f"""
                SELECT asx_code, company_name, sector,
                       price, return_1w, market_cap, volume, avg_volume_20d
                FROM screener.universe
                WHERE {liquid}
                  AND volume IS NOT NULL AND avg_volume_20d IS NOT NULL AND avg_volume_20d > 0
                  AND return_1w < 0
                  AND volume::float / avg_volume_20d >= 1.5
                ORDER BY volume::float / avg_volume_20d DESC NULLS LAST
                LIMIT :lim
            """), {"lim": limit})).mappings().all()

        except Exception as e:
            log.warning("volume_activity cap_tier=%s failed: %s", cap_tier, e)
            try:
                await db.rollback()
            except Exception:
                pass

    else:
        # Use pre-computed mover_snapshots — same source as the dashboard
        try:
            snap_date_row = (await db.execute(text("""
                SELECT MAX(snapshot_date) AS latest FROM market.mover_snapshots
            """))).mappings().one()
            snap_date = snap_date_row["latest"]
            if snap_date:
                rows = (await db.execute(text("""
                    SELECT snapshot_type, rank, asx_code, company_name, sector,
                           price, return_1w, market_cap, volume, avg_volume_20d
                    FROM market.mover_snapshots
                    WHERE snapshot_date = :d
                      AND snapshot_type IN ('ACTIVE', 'BUYING', 'SELLING')
                    ORDER BY snapshot_type, rank
                """), {"d": snap_date})).mappings().all()
                for r in rows:
                    st = r["snapshot_type"]
                    if   st == "ACTIVE":  active_rows.append(r)
                    elif st == "BUYING":  buying_rows.append(r)
                    elif st == "SELLING": selling_rows.append(r)
        except Exception as e:
            log.warning("volume_activity snapshot query failed: %s", e)

    result = VolumeActivityResponse(
        most_active  =[_to_active(r)       for r in active_rows],
        heavy_buying =[_to_vol_pressure(r) for r in buying_rows],
        heavy_selling=[_to_vol_pressure(r) for r in selling_rows],
        cap_tier=cap_tier,
    )
    ttl = MARKET_TTL if cap_tier else MARKET_TTL
    await cache_set(_key, result.model_dump(), ttl=ttl)
    return result


@router.get("/signals")
async def market_signals(
    limit: int = Query(15, ge=5, le=50),
    period: str = Query("1w", pattern="^(1d|1w|1m|3m|6m|1y|52w)$"),
    db: AsyncSession = Depends(get_db),
):
    """
    Market signals: stocks near their period high/low and volume surges.
    Uses market.period_metrics (pre-computed daily) for accurate H/L data.
    Falls back to a live CTE if the table has no data for today.
    """
    # Map period to column names, fallback lookback days, and minimum H/L range.
    # min_range filters out illiquid stocks with no real price movement in the period.
    col_map = {
        "1d":  ("high_1d",  "low_1d",  3,   0.003),  # 0.3% min daily swing
        "1w":  ("high_1w",  "low_1w",  7,   0.010),  # 1% min weekly swing
        "1m":  ("high_1m",  "low_1m",  35,  0.030),  # 3% min monthly swing
        "3m":  ("high_3m",  "low_3m",  100, 0.050),  # 5% min quarterly swing
        "6m":  ("high_6m",  "low_6m",  185, 0.080),
        "1y":  ("high_1y",  "low_1y",  365, 0.100),
        "52w": ("high_52w", "low_52w", 364, 0.100),
    }
    high_col, low_col, fallback_days, min_range = col_map[period]

    # Check if period_metrics has been populated for today
    has_table = (await db.execute(text("""
        SELECT 1 FROM market.period_metrics
        WHERE computed_date = CURRENT_DATE
        LIMIT 1
    """))).fetchone() is not None

    if has_table:
        # Fast path: read from pre-computed table
        period_join = f"""
            INNER JOIN market.period_metrics pm
                ON pm.asx_code = u.asx_code
               AND pm.computed_date = CURRENT_DATE
        """
        ph_expr = f"pm.{high_col}"
        pl_expr = f"pm.{low_col}"
    else:
        # Fallback: live CTE from daily_prices (used until first nightly run)
        period_join = f"""
            INNER JOIN (
                SELECT asx_code,
                       MAX(high) AS period_high,
                       MIN(low)  AS period_low
                FROM market.daily_prices
                WHERE time >= CURRENT_DATE - {fallback_days}
                  AND high > 0 AND low > 0
                GROUP BY asx_code
                HAVING COUNT(*) >= 1
            ) pm ON pm.asx_code = u.asx_code
        """
        ph_expr = "pm.period_high"
        pl_expr = "pm.period_low"

    # Near period high: within 5% of the high, stock must have had real movement,
    # and price must be strictly below the high (exclude exact 0.0% — no signal)
    high_rows = (await db.execute(text(f"""
        SELECT
            u.asx_code, u.company_name, u.sector, u.price, u.market_cap,
            {ph_expr} AS period_high,
            {pl_expr} AS period_low,
            u.volume, u.avg_volume_20d, u.return_1w, u.return_1m,
            ROUND(((u.price - {ph_expr}) / NULLIF({ph_expr}, 0) * 100)::numeric, 2) AS pct_from_high
        FROM screener.universe u
        {period_join}
        WHERE u.price > 0.10
          AND u.market_cap > 50
          AND {ph_expr} > 0
          AND {pl_expr} > 0
          AND {ph_expr} > {pl_expr} * (1 + {min_range})
          AND u.price >= {ph_expr} * 0.95
          AND u.price < {ph_expr} * 0.9999
        ORDER BY u.price / NULLIF({ph_expr}, 0) ASC
        LIMIT :lim
    """), {"lim": limit})).mappings().all()

    # Near period low: within 5% of the low, stock must have had real movement,
    # and price must be strictly above the low (exclude exact 0.0%)
    low_rows = (await db.execute(text(f"""
        SELECT
            u.asx_code, u.company_name, u.sector, u.price, u.market_cap,
            {ph_expr} AS period_high,
            {pl_expr} AS period_low,
            u.volume, u.avg_volume_20d, u.return_1w, u.return_1m,
            ROUND(((u.price - {pl_expr}) / NULLIF({pl_expr}, 0) * 100)::numeric, 2) AS pct_from_low
        FROM screener.universe u
        {period_join}
        WHERE u.price > 0.10
          AND u.market_cap > 50
          AND {ph_expr} > 0
          AND {pl_expr} > 0
          AND {ph_expr} > {pl_expr} * (1 + {min_range})
          AND u.price <= {pl_expr} * 1.05
          AND u.price > {pl_expr} * 1.0001
        ORDER BY u.price / NULLIF({pl_expr}, 0) DESC
        LIMIT :lim
    """), {"lim": limit})).mappings().all()

    # Volume surge (volume > 2× 20D average) — sorted by absolute volume DESC
    vol_rows = (await db.execute(text("""
        SELECT
            asx_code, company_name, sector, price, market_cap,
            volume, avg_volume_20d,
            return_1w, return_1m,
            ROUND((volume::numeric / NULLIF(avg_volume_20d, 0)), 1) AS vol_ratio
        FROM screener.universe
        WHERE price > 0.10
          AND market_cap > 50
          AND volume IS NOT NULL AND avg_volume_20d IS NOT NULL
          AND avg_volume_20d > 10000
          AND volume > avg_volume_20d * 2
        ORDER BY volume::numeric DESC
        LIMIT :lim
    """), {"lim": limit})).mappings().all()

    def _row(r, extra_key=None):
        d = {
            "asx_code":       r["asx_code"],
            "company_name":   r["company_name"],
            "sector":         r["sector"],
            "price":          float(r["price"]) if r["price"] is not None else None,
            "market_cap":     float(r["market_cap"]) if r["market_cap"] is not None else None,
            "period_high":    float(r["period_high"]) if r.get("period_high") is not None else None,
            "period_low":     float(r["period_low"]) if r.get("period_low") is not None else None,
            "volume":         int(r["volume"]) if r["volume"] is not None else None,
            "avg_volume_20d": int(r["avg_volume_20d"]) if r["avg_volume_20d"] is not None else None,
            "return_1w":      float(r["return_1w"]) if r["return_1w"] is not None else None,
            "return_1m":      float(r["return_1m"]) if r["return_1m"] is not None else None,
        }
        if extra_key and r[extra_key] is not None:
            d[extra_key] = float(r[extra_key])
        return d

    return {
        "near_period_high": [_row(r, "pct_from_high") for r in high_rows],
        "near_period_low":  [_row(r, "pct_from_low")  for r in low_rows],
        "volume_surge":     [_row(r, "vol_ratio")      for r in vol_rows],
    }


@router.get("/sectors", response_model=SectorsResponse)
async def market_sectors(db: AsyncSession = Depends(get_db)):
    """
    Per-GICS-sector aggregate stats — stock count, avg P/E, avg yield, avg 1Y return, market cap.
    Ordered by total market cap descending.
    """
    sql = text("""
        SELECT
            sector,
            COUNT(*)                                                              AS stock_count,
            AVG(pe_ratio) FILTER (WHERE pe_ratio > 0 AND pe_ratio < 100)         AS avg_pe,
            AVG(dividend_yield) FILTER (WHERE dividend_yield > 0)                AS avg_dividend_yield,
            AVG(return_1y) FILTER (WHERE return_1y IS NOT NULL)                  AS avg_return_1y,
            SUM(market_cap) / 1000.0                                             AS total_market_cap_bn
        FROM screener.universe
        WHERE status = 'active'
          AND sector IS NOT NULL
        GROUP BY sector
        ORDER BY total_market_cap_bn DESC NULLS LAST
    """)
    rows = (await db.execute(sql)).mappings().all()
    return SectorsResponse(
        sectors=[
            SectorStat(
                sector=r["sector"],
                stock_count=int(r["stock_count"]),
                avg_pe=float(r["avg_pe"]) if r["avg_pe"] is not None else None,
                avg_dividend_yield=float(r["avg_dividend_yield"]) if r["avg_dividend_yield"] is not None else None,
                avg_return_1y=float(r["avg_return_1y"]) if r["avg_return_1y"] is not None else None,
                total_market_cap_bn=float(r["total_market_cap_bn"]) if r["total_market_cap_bn"] is not None else None,
            )
            for r in rows
        ]
    )


@router.get("/dashboard", response_model=MarketDashboard)
async def market_dashboard(db: AsyncSession = Depends(get_db)):
    """
    All-in-one market overview sourced from pre-computed snapshot tables.
    Returns the latest available snapshot date (today or most recent prior day).
    """
    # Resolve latest snapshot date with actual data (skip zero-stock weekend snapshots)
    date_row = (await db.execute(text("""
        SELECT MAX(snapshot_date) AS latest
        FROM market.index_snapshots
        WHERE stock_count > 0
    """))).mappings().one()
    snap_date = date_row["latest"]

    if snap_date is None:
        from fastapi import HTTPException
        raise HTTPException(
            status_code=503,
            detail="Market snapshot not yet available. Run compute/engine/market_snapshot.py first."
        )

    # Index snapshots
    idx_rows = (await db.execute(text("""
        SELECT index_code, stock_count, gainers, losers, unchanged,
               avg_return_1w, total_market_cap_bn
        FROM market.index_snapshots
        WHERE snapshot_date = :d
    """), {"d": snap_date})).mappings().all()
    idx = {r["index_code"]: r for r in idx_rows}

    def _snap(code: str) -> IndexSnapshot:
        r = idx.get(code, {})
        return IndexSnapshot(
            stock_count=int(r.get("stock_count") or 0),
            gainers=int(r.get("gainers") or 0),
            losers=int(r.get("losers") or 0),
            unchanged=int(r.get("unchanged") or 0),
            avg_return_1w=float(r["avg_return_1w"]) if r.get("avg_return_1w") is not None else None,
            total_market_cap_bn=float(r["total_market_cap_bn"]) if r.get("total_market_cap_bn") is not None else None,
        )

    # Sector heatmap
    sector_rows = (await db.execute(text("""
        SELECT sector, stock_count, gainers, losers, avg_return_1w, total_market_cap_bn
        FROM market.sector_snapshots
        WHERE snapshot_date = :d
        ORDER BY total_market_cap_bn DESC NULLS LAST
    """), {"d": snap_date})).mappings().all()

    # Mover snapshots
    mover_rows = (await db.execute(text("""
        SELECT snapshot_type, rank, asx_code, company_name, sector,
               price, return_1w, market_cap, volume, avg_volume_20d, short_pct
        FROM market.mover_snapshots
        WHERE snapshot_date = :d
        ORDER BY snapshot_type, rank
    """), {"d": snap_date})).mappings().all()

    by_type: dict[str, list] = {"GAINER": [], "LOSER": [], "ACTIVE": [], "BUYING": [], "SELLING": []}
    for r in mover_rows:
        by_type.setdefault(r["snapshot_type"], []).append(r)

    def _dash(r) -> DashboardStock:
        return DashboardStock(
            asx_code=r["asx_code"],
            company_name=r["company_name"],
            sector=r["sector"],
            price=float(r["price"]) if r["price"] is not None else None,
            return_1w=float(r["return_1w"]) if r["return_1w"] is not None else None,
            market_cap=float(r["market_cap"]) if r["market_cap"] is not None else None,
        )

    def _vol_pressure(r) -> VolumePressureStock:
        vol = int(r["volume"]) if r["volume"] is not None else None
        avg = int(r["avg_volume_20d"]) if r["avg_volume_20d"] is not None else None
        ratio = (r["volume"] / r["avg_volume_20d"]) if (r["volume"] and r["avg_volume_20d"]) else None
        return VolumePressureStock(
            asx_code=r["asx_code"],
            company_name=r["company_name"],
            sector=r["sector"],
            price=float(r["price"]) if r["price"] is not None else None,
            return_1w=float(r["return_1w"]) if r["return_1w"] is not None else None,
            market_cap=float(r["market_cap"]) if r["market_cap"] is not None else None,
            volume=vol,
            avg_volume_20d=avg,
            volume_ratio=float(ratio) if ratio is not None else None,
        )

    # Ex-div snapshots
    exdiv_rows = (await db.execute(text("""
        SELECT asx_code, company_name,
               ex_div_date::text AS ex_div_date,
               pay_date::text    AS pay_date,
               dps_ttm, dividend_yield, franking_pct
        FROM market.exdiv_snapshots
        WHERE snapshot_date = :d
        ORDER BY ex_div_date ASC
    """), {"d": snap_date})).mappings().all()

    return MarketDashboard(
        asx200=_snap("ASX200"),
        asx300=_snap("ASX300"),
        sector_heatmap=[
            SectorHeatmapItem(
                sector=r["sector"],
                stock_count=int(r["stock_count"]),
                avg_return_1w=float(r["avg_return_1w"]) if r["avg_return_1w"] is not None else None,
                total_market_cap_bn=float(r["total_market_cap_bn"]) if r["total_market_cap_bn"] is not None else None,
            )
            for r in sector_rows
        ],
        top_gainers=[_dash(r) for r in by_type["GAINER"]],
        top_losers=[_dash(r) for r in by_type["LOSER"]],
        most_active=[
            ActiveStock(
                asx_code=r["asx_code"],
                company_name=r["company_name"],
                sector=r["sector"],
                price=float(r["price"]) if r["price"] is not None else None,
                return_1w=float(r["return_1w"]) if r["return_1w"] is not None else None,
                market_cap=float(r["market_cap"]) if r["market_cap"] is not None else None,
                volume=int(r["volume"]) if r["volume"] is not None else None,
                avg_volume_20d=int(r["avg_volume_20d"]) if r["avg_volume_20d"] is not None else None,
            )
            for r in by_type["ACTIVE"]
        ],
        heavy_buying=[_vol_pressure(r) for r in by_type["BUYING"]],
        heavy_selling=[_vol_pressure(r) for r in by_type["SELLING"]],
        upcoming_exdiv=[
            ExDivStock(
                asx_code=r["asx_code"],
                company_name=r["company_name"],
                ex_div_date=r["ex_div_date"],
                pay_date=r["pay_date"],
                dps_ttm=float(r["dps_ttm"]) if r["dps_ttm"] is not None else None,
                dividend_yield=float(r["dividend_yield"]) if r["dividend_yield"] is not None else None,
                franking_pct=float(r["franking_pct"]) if r["franking_pct"] is not None else None,
            )
            for r in exdiv_rows
        ],
        period="1w",
        universe_built_at=snap_date.isoformat() if snap_date else None,
    )


# ── Performance Heatmap ───────────────────────────────────────────────────────

import datetime as _dt

HEATMAP_TTL = 900  # 15 min — data only changes after nightly pipeline


def _fmt_day_label(d) -> str:
    """'Wed 21 May' from a date object."""
    if d is None:
        return ""
    if isinstance(d, str):
        d = _dt.date.fromisoformat(d)
    return d.strftime("%a %-d %b") if hasattr(d, "strftime") else str(d)


def _fmt_week_label(week_start) -> str:
    """'W/E 23 May' — the Friday of the given ISO week-start (Monday)."""
    if week_start is None:
        return ""
    if isinstance(week_start, str):
        week_start = _dt.date.fromisoformat(week_start)
    friday = week_start + _dt.timedelta(days=4)
    return friday.strftime("W/E %-d %b") if hasattr(friday, "strftime") else str(friday)


@router.get("/heatmap", response_model=HeatmapResponse)
async def market_heatmap(
    mode:    str           = Query("days",  pattern="^(days|weeks)$"),
    sector:  str | None    = Query(None),
    min_cap: float         = Query(0.0, ge=0, description="Min market cap AUD millions"),
    db: AsyncSession = Depends(get_db),
):
    """
    Rolling 5-period performance heatmap for all ASX stocks.

    mode=days  → last 5 trading days, each column = daily % change
    mode=weeks → last 5 calendar weeks, each column = weekly % change
                 (week close vs prior week close)

    Returns rows sorted by market_cap DESC so large-caps appear first.
    Cached for 15 minutes (HEATMAP_TTL).
    """
    _key = make_key("market", "heatmap", mode,
                    sector or "all", str(int(min_cap)))
    cached = await cache_get(_key)
    if cached:
        return HeatmapResponse(**cached)

    min_cap_raw = min_cap * 1_000_000   # convert AUD M → AUD raw

    # ── Build SQL ─────────────────────────────────────────────────────────────
    sector_clause = "AND u.sector = :sector" if sector else ""

    if mode == "days":
        sql = text(f"""
            WITH six_dates AS (
                SELECT DISTINCT DATE(time) AS td
                FROM market.daily_prices
                ORDER BY td DESC
                LIMIT 6
            ),
            prices AS (
                SELECT dp.asx_code, DATE(dp.time) AS td, dp.close
                FROM market.daily_prices dp
                INNER JOIN six_dates s ON DATE(dp.time) = s.td
                WHERE dp.close > 0
            ),
            lagged AS (
                SELECT asx_code, td, close,
                       LAG(close) OVER (PARTITION BY asx_code ORDER BY td) AS prev_close,
                       ROW_NUMBER() OVER (PARTITION BY asx_code ORDER BY td DESC) AS rn
                FROM prices
            )
            SELECT
                u.asx_code, u.company_name, u.sector, u.industry,
                u.price, u.market_cap,
                MAX(CASE WHEN l.rn=1 AND l.prev_close>0
                    THEN (l.close - l.prev_close) / l.prev_close END) AS p1,
                MAX(CASE WHEN l.rn=2 AND l.prev_close>0
                    THEN (l.close - l.prev_close) / l.prev_close END) AS p2,
                MAX(CASE WHEN l.rn=3 AND l.prev_close>0
                    THEN (l.close - l.prev_close) / l.prev_close END) AS p3,
                MAX(CASE WHEN l.rn=4 AND l.prev_close>0
                    THEN (l.close - l.prev_close) / l.prev_close END) AS p4,
                MAX(CASE WHEN l.rn=5 AND l.prev_close>0
                    THEN (l.close - l.prev_close) / l.prev_close END) AS p5
            FROM screener.universe u
            INNER JOIN lagged l ON u.asx_code = l.asx_code
            WHERE u.status = 'active'
              AND u.price > 0.05
              AND u.market_cap > :min_cap
              {sector_clause}
            GROUP BY u.asx_code, u.company_name, u.sector, u.industry,
                     u.price, u.market_cap
            ORDER BY u.market_cap DESC NULLS LAST
        """)

        labels_sql = text("""
            SELECT DISTINCT DATE(time) AS td
            FROM market.daily_prices
            ORDER BY td DESC
            LIMIT 5
        """)

    else:  # weeks
        sql = text(f"""
            WITH weekly AS (
                SELECT
                    asx_code,
                    DATE_TRUNC('week', time)::date AS week_start,
                    (ARRAY_AGG(close ORDER BY time DESC))[1] AS week_close
                FROM market.daily_prices
                WHERE time >= NOW() - INTERVAL '9 weeks'
                  AND close > 0
                GROUP BY asx_code, DATE_TRUNC('week', time)
            ),
            lagged AS (
                SELECT
                    asx_code, week_start, week_close,
                    LAG(week_close) OVER (PARTITION BY asx_code ORDER BY week_start)
                        AS prev_close,
                    ROW_NUMBER() OVER (PARTITION BY asx_code ORDER BY week_start DESC)
                        AS rn
                FROM weekly
            )
            SELECT
                u.asx_code, u.company_name, u.sector, u.industry,
                u.price, u.market_cap,
                MAX(CASE WHEN l.rn=1 AND l.prev_close>0
                    THEN (l.week_close - l.prev_close) / l.prev_close END) AS p1,
                MAX(CASE WHEN l.rn=2 AND l.prev_close>0
                    THEN (l.week_close - l.prev_close) / l.prev_close END) AS p2,
                MAX(CASE WHEN l.rn=3 AND l.prev_close>0
                    THEN (l.week_close - l.prev_close) / l.prev_close END) AS p3,
                MAX(CASE WHEN l.rn=4 AND l.prev_close>0
                    THEN (l.week_close - l.prev_close) / l.prev_close END) AS p4,
                MAX(CASE WHEN l.rn=5 AND l.prev_close>0
                    THEN (l.week_close - l.prev_close) / l.prev_close END) AS p5
            FROM screener.universe u
            INNER JOIN lagged l ON u.asx_code = l.asx_code
            WHERE u.status = 'active'
              AND u.price > 0.05
              AND u.market_cap > :min_cap
              {sector_clause}
              AND l.rn <= 5
            GROUP BY u.asx_code, u.company_name, u.sector, u.industry,
                     u.price, u.market_cap
            ORDER BY u.market_cap DESC NULLS LAST
        """)

        labels_sql = text("""
            SELECT DISTINCT DATE_TRUNC('week', time)::date AS week_start
            FROM market.daily_prices
            ORDER BY week_start DESC
            LIMIT 5
        """)

    # ── Execute ───────────────────────────────────────────────────────────────
    params: dict = {"min_cap": min_cap_raw}
    if sector:
        params["sector"] = sector

    rows_result  = (await db.execute(sql, params)).mappings().all()
    label_result = (await db.execute(labels_sql)).mappings().all()

    # ── Build labels ──────────────────────────────────────────────────────────
    if mode == "days":
        labels = [_fmt_day_label(r["td"]) for r in label_result]
    else:
        labels = [_fmt_week_label(r["week_start"]) for r in label_result]

    # Pad to 5 if needed
    while len(labels) < 5:
        labels.append("")

    # ── Build rows ────────────────────────────────────────────────────────────
    rows = [
        HeatmapRow(
            asx_code=r["asx_code"],
            company_name=r["company_name"],
            sector=r["sector"],
            industry=r["industry"],
            price=float(r["price"]) if r["price"] is not None else None,
            market_cap=float(r["market_cap"]) if r["market_cap"] is not None else None,
            p1=float(r["p1"]) if r["p1"] is not None else None,
            p2=float(r["p2"]) if r["p2"] is not None else None,
            p3=float(r["p3"]) if r["p3"] is not None else None,
            p4=float(r["p4"]) if r["p4"] is not None else None,
            p5=float(r["p5"]) if r["p5"] is not None else None,
        )
        for r in rows_result
    ]

    result = HeatmapResponse(rows=rows, labels=labels, mode=mode, total=len(rows))
    await cache_set(_key, result.model_dump(), ttl=HEATMAP_TTL)
    return result


# ── Heatmap Excel Export ──────────────────────────────────────────────────────

@router.get("/heatmap/export")
async def market_heatmap_export(
    mode:    str        = Query("days",  pattern="^(days|weeks)$"),
    sector:  str | None = Query(None),
    min_cap: float      = Query(0.0, ge=0),
    db: AsyncSession    = Depends(get_db),
):
    """
    Download the performance heatmap as a colour-coded Excel (.xlsx) file.
    Same query as /heatmap but streams an openpyxl workbook.
    """
    import io
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEP1
        from openpyxl.utils import get_column_letter
    except ImportError:
        from fastapi import HTTPException
        raise HTTPException(503, "openpyxl not installed — run: pip install openpyxl==3.1.5")

    # Re-use the JSON endpoint to get data (hits cache if warm)
    heatmap = await market_heatmap(mode=mode, sector=sector, min_cap=min_cap, db=db)

    # ── Colour map (hex fills) ─────────────────────────────────────────────────
    def _fill(hex_bg: str) -> PatternFill:
        return PatternFill("solid", fgColor=hex_bg)

    def _heat_fill(pct) -> PatternFill:
        if pct is None:
            return _fill("F3F4F6")   # gray-100
        if pct >=  0.05: return _fill("047857")   # emerald-700
        if pct >=  0.02: return _fill("10B981")   # emerald-500
        if pct >=  0.005: return _fill("A7F3D0")  # emerald-200
        if pct >= -0.005: return _fill("FDBA74")   # orange-300
        if pct >= -0.02: return _fill("FECACA")   # red-200
        if pct >= -0.05: return _fill("EF4444")   # red-500
        return _fill("B91C1C")                     # red-700

    def _heat_font(pct) -> Font:
        if pct is None:
            return Font(color="9CA3AF", size=9)
        if abs(pct) < 0.005:  # flat — dark text on orange
            return Font(color="7C2D12", size=9, bold=True)
        dark_text = abs(pct) < 0.02
        color = "065F46" if (pct >= 0.005 and dark_text) else \
                "991B1B" if (pct < -0.005 and dark_text) else "FFFFFF"
        return Font(color=color, size=9, bold=True)

    def _pct_str(v) -> str:
        if v is None:
            return "—"
        p = v * 100
        return ("+" if p >= 0 else "") + f"{p:.1f}%"

    def _cap_str(v) -> str:
        if v is None:
            return "—"
        m = v / 1_000_000
        if m >= 1000:
            return f"${m/1000:.1f}B"
        return f"${m:.0f}M"

    # ── Build workbook ────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"ASX Heatmap ({mode.capitalize()})"

    # Header row
    period_type = "Day" if mode == "days" else "Week"
    headers = [
        "ASX Code", "Company Name", "Sector", "Industry",
        "Price (AUD)", "Market Cap (AUD M)",
        *[f"Current{period_type}-{i+1}" for i in range(5)],
    ]

    # Add actual date labels in row 1 as a sub-header if available
    if any(heatmap.labels):
        sub_headers = ["", "", "", "", "", "", *heatmap.labels]
    else:
        sub_headers = None

    hdr_fill   = _fill("1E40AF")   # blue-800
    hdr_font   = Font(color="FFFFFF", bold=True, size=10)
    sub_fill   = _fill("3B82F6")   # blue-500
    sub_font   = Font(color="FFFFFF", size=9)
    center_al  = Alignment(horizontal="center", vertical="center", wrap_text=False)
    thin       = Side(style="thin", color="D1D5DB")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(headers)
    for col_idx, hdr in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill     = hdr_fill
        cell.font     = hdr_font
        cell.alignment = center_al
        cell.border   = border

    if sub_headers:
        ws.append(sub_headers)
        for col_idx, shdr in enumerate(sub_headers, start=1):
            cell = ws.cell(row=2, column=col_idx)
            cell.fill     = sub_fill
            cell.font     = sub_font
            cell.alignment = center_al
            cell.border   = border
        data_start_row = 3
    else:
        data_start_row = 2

    # Data rows
    for row in heatmap.rows:
        data = [
            row.asx_code,
            row.company_name,
            row.sector or "",
            row.industry or "",
            row.price,
            row.market_cap / 1_000_000 if row.market_cap else None,
            row.p1, row.p2, row.p3, row.p4, row.p5,
        ]
        ws.append(data)
        r = ws.max_row

        # Style meta columns (A–F)
        for c in range(1, 7):
            cell = ws.cell(row=r, column=c)
            cell.alignment = Alignment(vertical="center")
            cell.border    = border
            if c == 6 and cell.value is not None:  # Market Cap
                cell.number_format = '#,##0.0'

        # Style heat columns (G–K)
        for i, pct_val in enumerate([row.p1, row.p2, row.p3, row.p4, row.p5], start=7):
            cell = ws.cell(row=r, column=i)
            cell.value     = _pct_str(pct_val)
            cell.fill      = _heat_fill(pct_val)
            cell.font      = _heat_font(pct_val)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = border

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = [8, 38, 18, 22, 10, 16, 11, 11, 11, 11, 11]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Row heights
    ws.row_dimensions[1].height = 18
    if sub_headers:
        ws.row_dimensions[2].height = 14
    for r in range(data_start_row, ws.max_row + 1):
        ws.row_dimensions[r].height = 16

    # Freeze panes below headers, after ASX Code column
    ws.freeze_panes = f"C{data_start_row}"

    # ── Stream response ───────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    today = _dt.date.today().isoformat()
    filename = f"ASX_Heatmap_{mode}_{today}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
